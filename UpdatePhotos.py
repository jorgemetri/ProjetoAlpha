import pandas as pd
import re
import time
from bs4 import BeautifulSoup
import os
import shutil
import pandas as pd
import numpy as np
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from Automation import VerifyPaths
from Automation import logInProjetoAlpha,logIn
import openpyxl
from openpyxl.styles import Alignment

def VerifySheetName():
    # Define the file name
    file_name = 'raspagem.xlsx'
    
    # Get the current directory path
    current_directory = os.getcwd()
    
    # Create the full path to the file
    file_path = os.path.join(current_directory, file_name)
    
    # Check if the file exists
    return os.path.isfile(file_path)


def substituir_virgula_por_ponto(nome_arquivo, nome_aba, nome_coluna, novo_nome_arquivo):
    # Carrega o workbook e a worksheet específica
    wb = openpyxl.load_workbook(nome_arquivo)
    ws = wb[nome_aba]

    # Encontrar o índice da coluna pelo nome
    col_idx = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == nome_coluna:
            col_idx = col[0].column_letter
            break
    
    if not col_idx:
        raise ValueError(f"A coluna '{nome_coluna}' não foi encontrada na planilha.")
    
    # Iterar sobre as células na coluna especificada
    for cell in ws[col_idx]:
        if cell.row == 1:  # Pular o cabeçalho
            continue
        # Substituir vírgula por ponto se existir vírgula
        if ',' in str(cell.value):
            cell.value = str(cell.value).replace(',', '.')
        # Garantir que todos os valores sejam tratados como texto
        cell.value = f'="{cell.value}"' if cell.value else ''
        cell.alignment = Alignment(horizontal='left')

    # Salva o workbook atualizado
    wb.save(novo_nome_arquivo)
    print(f"Arquivo salvo como '{novo_nome_arquivo}'.")

def substituir_token(url, novo_token):
    padrao = r'token=[^&]*'
    nova_url = re.sub(padrao, f'token={novo_token}', url)
    return nova_url

def substituir_token_na_coluna( novo_token,arquivo_excel='raspagem.xlsx', coluna='Foto das Fazendas', arquivo_saida='raspagem.xlsx'):
    # Ler a planilha do Excel
  
   
    df = pd.read_excel(arquivo_excel, engine='openpyxl', dtype={'ID Fazendas': str})
    
    # Verificar se a coluna especificada existe
    if coluna not in df.columns:
        raise ValueError(f"A coluna '{coluna}' não existe no arquivo Excel.")
    
    # Função para substituir o token em uma URL
   
    
    # Aplicar a substituição na coluna especificada
    df[coluna] = df[coluna].apply(lambda x: substituir_token(x, novo_token) if isinstance(x, str) else x)
    
    # Salvar o DataFrame atualizado em um novo arquivo Excel
    df.to_excel(arquivo_excel, index=False)
    
    print(f"Arquivo salvo como {arquivo_saida}")


def TryGetImageUpdatePhotos(driver,attempts=200,delay=1):

    attempt =1
    contador = 1
    while attempt <= attempts:
        try:
            DataPage = BeautifulSoup(driver.page_source,'html.parser')
            panel = DataPage.find('question-answer')
            lista_fotos = panel.find_all('img',attrs={'class':'0 ir ir-middle'}) 
            if len(lista_fotos) == 0:
                raise Exception
            else:
                return lista_fotos[0]['src']
        except Exception as e:
            print(f'Tentando encontrar a imagem{'.'*contador}')
            contador+=1
            if contador == 5:
                contador = 1
        attempt+=1
        time.sleep(0.5)

def UpdatePhotosProjetoTerra(self,submain2_instance, download, target,chromedriver):
    


    if download == "" or target == "":
        submain2_instance.insert_text("Error! Por favor Preencha os campos adequadamente para antes de iniciar a automação.")
    elif not VerifySheetName():
        submain2_instance.insert_text('Error! Por favor coloque o nome da planilha de fotos como: raspagem.xlsx')
    else:
        options = Options()
        options.add_argument("--headless")

            #PARÂMETRO PARA MANTER WEBDRIVER ATUALIZADO
        servico = Service(ChromeDriverManager().install())

            #ENVIANDO PARÂMETROS PARA O DRIVER
        driver = webdriver.Chrome(options=options,service=servico)



        pathsWebdriver = chromedriver
        #downloadspath = str(input('Digite o caminho do download: '))
        #targetpath = str(input('Digite o caminho onde o programa está sendo executado: '))
        #pathsWebdriver = arg3
        downloadspath = download
        targetpath = target
        driver=logIn(submain2_instance,pathsWebdriver,driver)
        lista_fotos_final = []
    
        for i in range(1,2):
            #Iterate over each element on table
            
            driver.get(f'https://survey123.arcgis.com/surveys/7c7eb1d6f2b74421a50e01a2d9eb2667/data?extent=-80.6008,-21.9941,-13.8039,-9.7200&objectIds={(i)}')
            #Trying get element


            url=TryGetImageUpdatePhotos(driver)
            print(f'Url obtida para ser atualizada!')
            match = re.search(r'token=([^&]+)',url)
            if match:
                token = match.group(1)
                print('Substituindo novo token na tabela!')
                substituir_token_na_coluna(token)
            else:
                print("Token not found")

def UpdatePhotosProjetoTerra1(self,submain2_instance, download, target,chromedriver):
    


    if download == "" or target == "":
        submain2_instance.insert_text("Error! Por favor Preencha os campos adequadamente para antes de iniciar a automação.")
    elif not VerifySheetName():
        submain2_instance.insert_text('Error! Por favor coloque o nome da planilha de fotos como: raspagem.xlsx')
    else:
        options = Options()
        options.add_argument("--headless")

        # Definindo o caminho diretamente para o ChromeDriver
        try:
            service = Service(chromedriver)
            driver = webdriver.Chrome(service=service, options=options)
        except OSError as e:
            print(f"Erro ao iniciar o ChromeDriver: {e}")
            return

        downloadspath = download
        targetpath = target

        if len(VerifyPaths()) > 0:
            pass
        else:
            lines = [f'{download}\n', f'{target}\n']
            with open('paths.txt', 'w') as file:
                file.writelines(lines)



        pathsWebdriver = chromedriver
        #downloadspath = str(input('Digite o caminho do download: '))
        #targetpath = str(input('Digite o caminho onde o programa está sendo executado: '))
        #pathsWebdriver = arg3
        downloadspath = download
        targetpath = target
        driver=logIn(submain2_instance,pathsWebdriver,driver)
        lista_fotos_final = []
    
        for i in range(1,2):
            #Iterate over each element on table
            
            driver.get(f'https://survey123.arcgis.com/surveys/7c7eb1d6f2b74421a50e01a2d9eb2667/data?extent=-80.6008,-21.9941,-13.8039,-9.7200&objectIds={(i)}')
            #Trying get element


            url=TryGetImageUpdatePhotos(driver)
            print(f'Url obtida para ser atualizada!')
            match = re.search(r'token=([^&]+)',url)
            if match:
                token = match.group(1)
                print('Substituindo novo token na tabela!')
                substituir_token_na_coluna(token)
            else:
                print("Token not found")

def replace_commas_with_periods_in_excel(file_path = 'raspagem.xlsx', column_name='ID Fazendas'):
    """
    Replace commas with periods in a specified column of a DataFrame loaded from an Excel file,
    and save the updated DataFrame back to the same Excel file.

    Parameters:
    file_path (str): The path to the Excel file.
    column_name (str): The name of the column to process.
    """
    # Load the DataFrame from the Excel file
    df = pd.read_excel(file_path)

    # Replace commas with periods only if the row contains a comma
    df[column_name] = df[column_name].astype(str).apply(lambda x: x.replace(',', '.') if ',' in x else x)

    # Ensure values without commas remain as strings to avoid float conversion
    df[column_name] = df[column_name].apply(lambda x: x.rstrip('.0') if '.' in x and x.endswith('.0') else x)

    # Save the updated DataFrame back to the same Excel file
    df.to_excel(file_path, index=False)

def UpdatePhotosAlphaProject(self,submain2_instance, download, target,chromedriver):

    


    if download == "" or target == "":
        submain2_instance.insert_text("Error! Por favor Preencha os campos adequadamente para antes de iniciar a automação.")
    elif not VerifySheetName():
        submain2_instance.insert_text("Error! Por favor coloque o nome da planilha de imagens como : raspagem.xlsx.")
    else:
        options = Options()
        options.add_argument("--headless")

        #PARÂMETRO PARA MANTER WEBDRIVER ATUALIZADO
        servico = Service(ChromeDriverManager().install())

        #ENVIANDO PARÂMETROS PARA O DRIVER
        driver = webdriver.Chrome(options=options,service=servico)



        pathsWebdriver = chromedriver
        #downloadspath = str(input('Digite o caminho do download: '))
        #targetpath = str(input('Digite o caminho onde o programa está sendo executado: '))
        #pathsWebdriver = arg3
        downloadspath = download
        targetpath = target
        driver=logInProjetoAlpha(pathsWebdriver,driver)

        lista_fotos_final = []
    
        for i in range(1,2):
            #Iterate over each element on table
            
            driver.get (f'https://survey123.arcgis.com/surveys/75bc7c0c6e674430b13433169665526f/data?extent=63.1750,-57.1147,51.9250,81.8793&objectIds={(i)}')
            #Trying get element

        
            url=TryGetImageUpdatePhotos(driver)
            match = re.search(r'token=([^&]+)',url)
            if match:
                token = match.group(1)
                substituir_token_na_coluna(token)
            else:
                print("Token not found")
            substituir_virgula_por_ponto('raspagem.xlsx', 'Sheet1', 'ID Fazendas', 'raspagem.xlsx')
def UpdatePhotosAlphaProject1(self,submain2_instance, download, target,chromedriver):
    


    if download == "" or target == "":
        submain2_instance.insert_text("Error! Por favor Preencha os campos adequadamente para antes de iniciar a automação.")
    elif not VerifySheetName():
        submain2_instance.insert_text('Error! Por favor coloque o nome da planilha de fotos como: raspagem.xlsx')
    else:
        options = Options()
        options.add_argument("--headless")

        # Definindo o caminho diretamente para o ChromeDriver
        try:
            service = Service(chromedriver)
            driver = webdriver.Chrome(service=service, options=options)
        except OSError as e:
            print(f"Erro ao iniciar o ChromeDriver: {e}")
            return

        downloadspath = download
        targetpath = target

        if len(VerifyPaths()) > 0:
            pass
        else:
            lines = [f'{download}\n', f'{target}\n']
            with open('paths.txt', 'w') as file:
                file.writelines(lines)



        pathsWebdriver = chromedriver
        #downloadspath = str(input('Digite o caminho do download: '))
        #targetpath = str(input('Digite o caminho onde o programa está sendo executado: '))
        #pathsWebdriver = arg3
        downloadspath = download
        targetpath = target
        driver=logIn(submain2_instance,pathsWebdriver,driver)
        lista_fotos_final = []
    
        for i in range(1,2):
            #Iterate over each element on table
            
            driver.get(f'https://survey123.arcgis.com/surveys/75bc7c0c6e674430b13433169665526f/data?extent=63.1750,-57.1147,51.9250,81.8793&objectIds={(i)}')
            #Trying get element
            

            url=TryGetImageUpdatePhotos(driver)
            print(url)
            print(f'Url obtida para ser atualizada!')
            match = re.search(r'token=([^&]+)',url)
            if match:
                token = match.group(1)
                print('Substituindo novo token na tabela!')
                substituir_token_na_coluna(token)
            else:
                print("Token not found")
