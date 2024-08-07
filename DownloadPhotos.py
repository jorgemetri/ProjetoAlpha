import os
import shutil
import pandas as pd
import numpy as np
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException
from selenium.webdriver import ActionChains
import openpyxl
from openpyxl.utils import get_column_letter
import re
from geopy.geocoders import Nominatim



def FecharJanela(driver):
    actions = ActionChains(driver)
    actions.send_keys(Keys.ESCAPE)
    actions.perform()

#Old GetCityB, Before Alpha Project-----------------------------------------------------------------------------------------------
# def GetCityB(driver):
    
#     tentativas = 0

#     while tentativas <= 4:
#         try:
#             img_element = driver.find_elements(By.CSS_SELECTOR, 'img.ir.ir-middle')
#             img_element[tentativas].click()
#             print('Clickou na Image!!!')
#             time.sleep(1.5)
#             elements = driver.find_elements(By.CSS_SELECTOR, 'div.item > p.main-info')
#             str=elements[len(elements)-1].text
#             array = str.split(', ')
#             print(f'Imagens geradas:{array}')
#             print(f'Tentivas pegar cidade na imagem {tentativas}')
#             return  array[1]
#         except Exception:
#             raise Exception('Nao encontrou a cidade!')
#             FecharJanela(driver)
#         tentativas= tentativas+1

#     return ''
#-----------------------------------------------------------------------------------------------------------------------------------------
#New GetCityB, After Alpha Project-------------------------------------------------------------------------------------------------------
def GetCityB(driver):
    
    tentativas = 0
    print('Tentar Pegar cidade------------------------------------------------------------------')
    while tentativas <= 4:
        try:
            img_element = driver.find_elements(By.CSS_SELECTOR, 'img.ir.ir-middle')
            img_element[tentativas].click()
            print('Clickou na Image!!!')
            time.sleep(1)
            elements = driver.find_elements(By.CSS_SELECTOR, 'div.item > p.main-info')
            str=elements[len(elements)-1].text
            array = str.split(', ')
            print(f'Imagens geradas:{array}')
            print(f'Tentivas pegar cidade na imagem {tentativas}')
            return  array[1]
        except Exception:
            print('Nao encontrou a cidade!')
            FecharJanela(driver)
            
        tentativas= tentativas+1
    cid=find_city_from_page_source(driver)#New Line Added----------------------------------------------------------------------------------------------
    if cid != '':
        return cid
    return ''
#-----------------------------------------------------------------------------------------------------------------------------------------
def save_last_index(index):
    """
    My text:
    This function will to save last index from PegarElemento function on for loop. This index will be save in
    txt file. So every time that occurs some error, PegarElemento know what was the last index in for loop.
    """
    """
    Grammar fixed:
    This function will save the last index from the PegarElemento function in a for loop. This index will 
    be saved in a text file. So every time an error occurs, PegarElemento will know what the last index in the for loop was.
    """
    filename = 'lastindex.txt'
    try:
        with open(filename, 'w') as file:
            file.write(str(index))
        print(f'O valor {index} foi salvo no arquivo {filename}.')
    except Exception as e:
        print(f'Ocorreu um erro ao salvar o valor: {e}')

def read_last_index():
    filename = 'lastindex.txt'
    try:
        with open(filename, 'r') as file:
            index = int(file.read().strip())
        return index
    except FileNotFoundError:
        print(f"O arquivo {filename} não foi encontrado.")
        return None
    except ValueError:
        print(f"O arquivo {filename} não contém um valor inteiro válido.")
        return None
    except Exception as e:
        print(f"Ocorreu um erro ao ler o valor: {e}")
        return None

#DeletarExcelExistente---------------------------------------------------------------------------------------------------------------
def delete_excel_file(file_path):
    """
    Deletes the specified Excel file.

    Parameters:
    file_path (str): The path to the Excel file to be deleted.

    Returns:
    str: A message indicating whether the file was deleted successfully or not.
    """
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            return f"File '{file_path}' has been deleted successfully."
        except Exception as e:
            return f"An error occurred while deleting the file: {e}"
    else:
        return f"File '{file_path}' does not exist."


#Funções Modificadas-------------------------------------------------------------------------------------

def TentarPegarElementoModificado(driver, retries=35, delay=1):
    """
    This function will try to get the element every time this task occurs, if possible.
    An error may be raised, so every time these errors occur, this function will be called again.
    It will retry up to 'retries' times with a delay of 'delay' seconds between attempts.
    """
    attempt = 0
    lista_dados_parcial = []

    while attempt < retries:
        try:
            # Parse the page source with BeautifulSoup
            DataPage = BeautifulSoup(driver.page_source, 'html.parser')
           

            # Encontrar o elemento <p> que contém um <b> com o texto 'ID Fazenda'
            element = DataPage.find('p', string=lambda t: t and 'ID Fazenda' in t)

            if element is not None:
                # Encontrar o elemento <p> irmão com a classe 'answer'
                answer_element = element.find_next('p', class_='answer')
                
                if answer_element:
                    answer_text = answer_element.get_text()
                    print(f'Fazenda:{answer_text}')
                else:
                    print("Elemento irmão com a classe 'answer' não encontrado.")
            else:
                print("Elemento contendo <b>ID Fazenda</b> não encontrado.")
                

            #This code is responsible for getting and clicking on the image
            if answer_element != "":

                _cidade = GetCityB(driver)
                if _cidade == '':
                    raise Exception(f"Failed to find the element after {retries} attempts")
                # Close the current window
                print(answer_text)
                FecharJanela(driver)
                return [_cidade,answer_text]
            
        except AttributeError as e1:
            print(f"AttributeError: {e1}")
        
        except NoSuchElementException as e2:
            print(f"NoSuchElementException: {e2}")
        
        except ElementNotInteractableException as e3:
            print(f"ElementNotInteractableException: {e3}")
        
        except ElementClickInterceptedException as e4:
            print(f"ElementClickInterceptedException: {e4}")
        
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

        
    
        attempt += 1
        print(f"Retrying... ({attempt}/{retries})")
        time.sleep(delay)
    
    raise Exception(f"Failed to find the element after {retries} attempts")

def verificar_arquivo_excel(arquivo_excel):
    """
    Verifica se um arquivo Excel existe.

    Parâmetros:
    arquivo_excel (str): Caminho do arquivo Excel.

    Retorna:
    bool: True se o arquivo existir, False caso contrário.
    """
    return os.path.isfile(arquivo_excel)

def adicionar_dados_dataframe(arquivo_excel, dataframe):
    try:
        # Tenta abrir o arquivo Excel
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active
        
        # Encontra a última linha com dados
        ultima_linha = sheet.max_row
        
        # Verifica se as colunas do DataFrame correspondem às colunas do arquivo Excel
        colunas_excel = [cell.value for cell in sheet[1]]
        colunas_dataframe = list(dataframe.columns)
        
        if colunas_excel != colunas_dataframe:
            raise ValueError("As colunas do DataFrame não correspondem às colunas do arquivo Excel.")
        
        # Adiciona os novos dados do DataFrame ao arquivo Excel
        for index, row in dataframe.iterrows():
            for col, valor in enumerate(row, start=1):
                sheet.cell(row=ultima_linha + 1 + index, column=col, value=valor)
        
        # Salva o arquivo Excel
        workbook.save(arquivo_excel)
        print("Dados adicionados com sucesso.")
    except Exception as e:
        print(f"Ocorreu um erro ao tentar abrir ou modificar o arquivo Excel: {e}")


def adicionar_dados_ultima_linha(arquivo_excel, dados):
    try:
        # Tenta abrir o arquivo Excel
        workbook = openpyxl.load_workbook(arquivo_excel)
        sheet = workbook.active
        
        # Encontra a última linha com dados
        ultima_linha = sheet.max_row
        
        # Adiciona os novos dados na próxima linha
        for col, valor in enumerate(dados, start=1):
            sheet.cell(row=ultima_linha + 1, column=col, value=valor)
        
        # Salva o arquivo Excel
        workbook.save(arquivo_excel)
        print("Dados adicionados com sucesso.")
    except Exception as e:
        print(f"Ocorreu um erro ao tentar abrir ou modificar o arquivo Excel: {e}")


def AddArrayDataframe(df_vazio,array_dados):
# Adicionar cada linha do array ao DataFrame
    for i, row in enumerate(array_dados):
        df_vazio.loc[i] = row


#Complemetary Function Alpha Project--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def extract_lat_long(text):
    lat_long_regex = r"Lat:\s*([-+]?\d*\.\d+|\d+)\s*Long:\s*([-+]?\d*\.\d+|\d+)"
    match = re.search(lat_long_regex, text)
    if match:
        latitude, longitude = match.groups()
        return float(latitude), float(longitude)
    return None, None


def get_city_name(latitude, longitude):
    geolocator = Nominatim(user_agent="my_unique_app_name")
    geocode = RateLimiter(geolocator.reverse, min_delay_seconds=2)
    location = geocode((latitude, longitude), exactly_one=True)
    address = location.raw['address']
    
    city = address.get('city', '')
    if not city:
        city = address.get('town', '')
    if not city:
        city = address.get('village', '')
    
    return city



def find_city_from_page_source(driver):
    try:
        dados_pg = BeautifulSoup(driver.page_source, 'html.parser')
        painel = dados_pg.find_all('p', 'answer')

        for element in painel:
            if 'Lat:' in element.text and 'Long:' in element.text:
                latitude, longitude = extract_lat_long(element.text)
                if latitude is not None and longitude is not None:
                    city = get_city_name(latitude, longitude)
                    return city

    except Exception as e3:
        print(f'Exception Ocurred in function: find_city_from_page_source:{e3}!')


#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def PegarElementoModificado(driver):
    lista_fotos_final = []
    lista_contador_final = []

    cidades_support = pd.DataFrame(columns=["Cidade", "ID Fazendas"])
    _Fotos = pd.DataFrame(columns=["Foto das Fazendas"])
    _contador = pd.DataFrame(columns=['Contador'])
    raspagem = pd.DataFrame(columns=['Contador','Cidade', 'ID Fazendas', 'Foto das Fazendas'])
    if verificar_arquivo_excel('raspagem.xlsx'):
        delete_excel_file('raspagem.xlsx')
        raspagem.to_excel('raspagem.xlsx', index=False)
    else:
        #Caso o arquivo excel nao exista crie ele
        raspagem.to_excel('raspagem.xlsx', index=False)
    totalValues = getTotalValues(driver)
    totalValues= totalValues+1

    for i in range(1, totalValues):
            # Acessar a página
            index = i
            driver.get(f'https://survey123.arcgis.com/surveys/7c7eb1d6f2b74421a50e01a2d9eb2667/data?extent=-80.6008,-21.9941,-13.8039,-9.7200&objectIds={(index)}')
            # Tentar obter o elemento
            try:
                _cidade, answer_text = TentarPegarElementoModificadoNovo(driver, retries=45)
                print('TryGetElement complete: Element its avaible for to get!')
            except Exception as e:
                print(f'Error during TryGetElement: {e}')
                continue

                # Obter dados da página
            DataPage = BeautifulSoup(driver.page_source, 'html.parser')
            panel = DataPage.find('question-answer')

            lista_fotos = panel.find_all('img', attrs={'class': '0 ir ir-middle'})
                
            for foto in lista_fotos:
                data_city = {_cidade: answer_text}
                    #print(data_city)
                cidades_df = pd.DataFrame(list(data_city.items()), columns=["Cidade", "ID Fazendas"])
                    #print(data_city.items())
                cidades_support = pd.concat([cidades_support, cidades_df], ignore_index=True)
                    #print(cidades_support)
                lista_fotos_final.append(foto['src'])
                lista_contador_final.append(index)


            print(len(lista_fotos_final))

                # Criar a planilha do Excel toda vez que o programa rodar

            AddArrayDataframe(_Fotos, lista_fotos_final)
            AddArrayDataframe(_contador,lista_contador_final)
                #print(_Fotos)
            print('Fotos dataframe:',_Fotos)
                #Criando o dataframe do contador-------------------------------------------------------------------------------
            contador = {index: answer_text}
            contador_df = pd.DataFrame(list(contador.items()), columns=["Contador", "ID Fazendas"])

            final_Data = cidades_support.join(_Fotos)
            final_Data = final_Data.join(_contador)
            print('DataFrame Final')
        
            final_Data = final_Data[['Contador', 'Cidade', 'ID Fazendas', 'Foto das Fazendas']]
            print(final_Data)
            adicionar_dados_dataframe('raspagem.xlsx', final_Data)

                # Esvaziar DataFrames
            contador_df = contador_df.iloc[0:0]
            cidades_support = cidades_support.iloc[0:0]
            _Fotos = _Fotos.iloc[0:0]
            final_Data=final_Data.iloc[0:0]
            lista_contador_final = []
            lista_fotos = []
            lista_fotos_final = []
#-----------------------------------------------------------------------------------------------------------------------
def CorrigirLinhasVazias(driver,df):
    lista_fotos_final = []
    lista_contador_final = []
    valores_unicos = df['Contador'].unique()
    cidades_support = pd.DataFrame(columns=["Cidade", "ID Fazendas"])
    _Fotos = pd.DataFrame(columns=["Foto das Fazendas"])
    _contador = pd.DataFrame(columns=['Contador'])

    for i in range(0, len(valores_unicos)):
            # Acessar a página
        index = int(valores_unicos[i])
        driver.get(f'https://survey123.arcgis.com/surveys/7c7eb1d6f2b74421a50e01a2d9eb2667/data?extent=-80.6008,-21.9941,-13.8039,-9.7200&objectIds={(index)}')
        # Tentar obter o elemento
        try:
            _cidade, answer_text = TentarPegarElementoModificadoNovo(driver, retries=45)
            print('TryGetElement complete: Element its avaible for to get!')
        except Exception as e:
            print(f'Error during TryGetElement: {e}')
            continue

            # Obter dados da página
        DataPage = BeautifulSoup(driver.page_source, 'html.parser')
        panel = DataPage.find('question-answer')

        lista_fotos = panel.find_all('img', attrs={'class': '0 ir ir-middle'})
            
        for foto in lista_fotos:
            data_city = {_cidade: answer_text}
                #print(data_city)
            cidades_df = pd.DataFrame(list(data_city.items()), columns=["Cidade", "ID Fazendas"])
                #print(data_city.items())
            cidades_support = pd.concat([cidades_support, cidades_df], ignore_index=True)
                #print(cidades_support)
            lista_fotos_final.append(foto['src'])
            lista_contador_final.append(index)


        print(len(lista_fotos_final))

            # Criar a planilha do Excel toda vez que o programa rodar

        AddArrayDataframe(_Fotos, lista_fotos_final)
        AddArrayDataframe(_contador,lista_contador_final)
            #print(_Fotos)
        print('Fotos dataframe:',_Fotos)
            #Criando o dataframe do contador-------------------------------------------------------------------------------
        contador = {index: answer_text}
        contador_df = pd.DataFrame(list(contador.items()), columns=["Contador", "ID Fazendas"])

        final_Data = cidades_support.join(_Fotos)
        final_Data = final_Data.join(_contador)
        print('DataFrame Final')
    
        final_Data = final_Data[['Contador', 'Cidade', 'ID Fazendas', 'Foto das Fazendas']]
        print(final_Data)
        adicionar_dados_dataframe('raspagem.xlsx', final_Data)

            # Esvaziar DataFrames
        contador_df = contador_df.iloc[0:0]
        cidades_support = cidades_support.iloc[0:0]
        _Fotos = _Fotos.iloc[0:0]
        final_Data=final_Data.iloc[0:0]
        lista_contador_final = []
        lista_fotos = []
    





#--------------------------------------------------------------------------------------------------------------------------------
def move_specific_file_Alpha_Project(path_download,destination_folder):
    """
    Move the last downloaded .xlsx file containing 'S123' in the filename from the Downloads folder to a specified destination folder.
    Overwrites the file if a file with the same name exists at the destination.
    
    Args:
    destination_folder (str): The path to the destination folder where the file should be moved.
    """
    # Formating Path from download and destinantion
    path_download = path_download.replace('/','//')
    destination_folder = destination_folder.replace('/','//')
    # Set the path to your Downloads folder
    downloads_folder = path_download
    # Ensure the destination folder exists, if not, create it
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    files =[]
    attempt = 120
    attempts=1
    while len(files) <= 0 and attempts <= attempt:
        
        files = [os.path.join(downloads_folder, file) for file in os.listdir(downloads_folder) if os.path.isfile(os.path.join(downloads_folder, file)) and file.endswith('.xlsx') and 'S123' in file]

        # Sort files by modification time in descending order
        files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        # Move the most recently downloaded file that meets the criteria
        if files:
            last_downloaded_file = files[0]
            new_file_path = os.path.join(destination_folder, os.path.basename(last_downloaded_file))

            # Check if there is a file with the same name in the destination folder and delete it if exists
            if os.path.exists(new_file_path):
                os.remove(new_file_path)
                print(f"Existing file with the same name was found and has been removed: {new_file_path}")

            shutil.move(last_downloaded_file, new_file_path)
            print(f"The last downloaded file has been moved to: {new_file_path}")
            return
        else:
            print("No suitable files found in the Downloads folder.")



        time.sleep(1)
        print(f'You dont get succed in find your folder. Attempt {attempts}')
        attempts+=1
        #Try get this folder Again


def TryDownloadSheetAlphaProject(driver,attempt=50,delay=1):
    attempts =1
    while attempts <= attempt:

        try:
            #Click on Exportar
            driver.find_element(By.XPATH,'/html/body/s123-app/div/main/ng-component/div/ng-component/div/div[1]/header-menubar/div/print-util-menus/div/div[2]/a').click()
            time.sleep(2)
            driver.find_element(By.XPATH, '/html/body/s123-app/div/main/ng-component/div/ng-component/div/div[1]/header-menubar/div/print-util-menus/div/div[2]/ul/li[4]/a/div/span').click()
            print('You get Succed and Download Sheet')
            return True

        except Exception as e:
            print('Try get element and dont get succed, Try Again {i}!')
        
        try:
            time.sleep(2)
            driver.find_element(By.XPATH,'/html/body/div[3]/div/div/div/div[1]/span[2]').click()
            print('Clicked on x in login stage!')
        except Exception as e:
            print('This element dont exist on page!(element close)')
            
        time.sleep(1)
        attempts+=1
#GetTotalValuesAlphaProject------------------------------------------------------------------------------------------------------
def getTotalValues(driver):
    element=driver.find_element(By.XPATH,'/html/body/s123-app/div/main/ng-component/div/ng-component/div/div[1]/header-menubar/div/feature-count/div').text
    return int(element.split('/')[0])
#PegarElementoModificadoAlphaProjetctNewVewsion-----------------------------------------------------------------------------------------------------------
def PegarElementoModificadoAlphaProject(driver):
    lista_fotos_final = []
    lista_contador_final = []

    cidades_support = pd.DataFrame(columns=["Cidade", "ID Fazendas"])
    _Fotos = pd.DataFrame(columns=["Foto das Fazendas"])
    _contador = pd.DataFrame(columns=['Contador'])
    raspagem = pd.DataFrame(columns=['Contador','Cidade', 'ID Fazendas', 'Foto das Fazendas'])
    if verificar_arquivo_excel('raspagem.xlsx'):
        delete_excel_file('raspagem.xlsx')
        raspagem.to_excel('raspagem.xlsx', index=False)
    else:
        #Caso o arquivo excel nao exista crie ele
        raspagem.to_excel('raspagem.xlsx', index=False)
    totalValues = getTotalValues(driver)
    totalValues= totalValues+1

    for i in range(7, totalValues+6):
            # Acessar a página
            index = i
            driver.get (f'https://survey123.arcgis.com/surveys/75bc7c0c6e674430b13433169665526f/data?extent=63.1750,-57.1147,51.9250,81.8793&objectIds={(index)}')
            # Tentar obter o elemento
            try:
                _cidade,answer_text=TentarPegarElementoModificadoNovo(driver,retries=35)
                print('TryGetElement complete: Element its avaible for to get !')
            except Exception as e:
                print(f'Error during TryGetElement: {e}')
                continue

                # Obter dados da página
            DataPage = BeautifulSoup(driver.page_source, 'html.parser')
            panel = DataPage.find('question-answer')

            lista_fotos = panel.find_all('img', attrs={'class': '0 ir ir-middle'})
                
            for foto in lista_fotos:
                data_city = {_cidade: answer_text}
                    #print(data_city)
                cidades_df = pd.DataFrame(list(data_city.items()), columns=["Cidade", "ID Fazendas"])
                    #print(data_city.items())
                cidades_support = pd.concat([cidades_support, cidades_df], ignore_index=True)
                    #print(cidades_support)
                lista_fotos_final.append(foto['src'])
                lista_contador_final.append(index)


            print(len(lista_fotos_final))

                # Criar a planilha do Excel toda vez que o programa rodar

            AddArrayDataframe(_Fotos, lista_fotos_final)
            AddArrayDataframe(_contador,lista_contador_final)
                #print(_Fotos)
            print('Fotos dataframe:',_Fotos)
                #Criando o dataframe do contador-------------------------------------------------------------------------------
            contador = {index: answer_text}
            contador_df = pd.DataFrame(list(contador.items()), columns=["Contador", "ID Fazendas"])

            final_Data = cidades_support.join(_Fotos)
            final_Data = final_Data.join(_contador)
            print('DataFrame Final')
        
            final_Data = final_Data[['Contador', 'Cidade', 'ID Fazendas', 'Foto das Fazendas']]
            print(final_Data)
            adicionar_dados_dataframe('raspagem.xlsx', final_Data)

                # Esvaziar DataFrames
            contador_df = contador_df.iloc[0:0]
            cidades_support = cidades_support.iloc[0:0]
            _Fotos = _Fotos.iloc[0:0]
            final_Data=final_Data.iloc[0:0]
            lista_contador_final = []
            lista_fotos = []
            lista_fotos_final = []

#PegarElementoModificadoAlphaProject-----------------------------------------------------------------------------------------------------------------------
# def PegarElementoModificadoAlphaProject(driver):
    # lista_fotos_final = []
    # cidades_support = pd.DataFrame()
    
    # # #Start i with last index
    # raspagem = pd.DataFrame(columns=['Cidade', 'ID Fazendas', 'Foto das Fazendas'])

    # # # Salvar o DataFrame em um arquivo Excel
    # if verificar_arquivo_excel('raspagem.xlsx'):
    #     delete_excel_file('raspagem.xlsx')
    #     raspagem.to_excel('raspagem.xlsx', index=False)
    # else:
    #     #Caso o arquivo excel nao exista crie ele
    #     raspagem.to_excel('raspagem.xlsx', index=False)
        
    # lastindex = read_last_index()
    # _Fotos = pd.DataFrame( columns=["Foto das Fazendas"])
    # totalValues = getTotalValues(driver)
    # totalValues= totalValues+1
    # for i in range(1,totalValues):
    #     #Iterate over each element on table
        
    #     driver.get (f'https://survey123.arcgis.com/surveys/75bc7c0c6e674430b13433169665526f/data?extent=63.1750,-57.1147,51.9250,81.8793&objectIds={(i)}')
    #     #Trying get element

    
    #     _cidade,answer_text=TentarPegarElementoModificadoNovo(driver,retries=35)
    #     print('TryGetElement complete: Element its avaible for to get !')
    #     #Get data now
    #     DataPage = BeautifulSoup(driver.page_source,'html.parser')
    #     panel = DataPage.find('question-answer')
        
        
    #     save_last_index(i)
    #     lista_fotos = panel.find_all('img',attrs={'class':'0 ir ir-middle'}) 
         
    #     for foto in lista_fotos:
    #         data_city={_cidade:answer_text}
    #         print(data_city)
    #         cidades_df = pd.DataFrame(list(data_city.items()),columns=["Cidade", "ID Fazendas"])
    #         print(data_city.items())
    #         cidades_support = pd.concat([cidades_support, cidades_df], ignore_index=True)
    #         print(cidades_support)
    #         lista_fotos_final.append(foto['src'])
    #     print(lista_fotos_final)
        

    #     #Criando a planilha do excel toda vez que o programa rodar
    #     AddArrayDataframe(_Fotos,lista_fotos_final)
    #     print('Fotos adicionadas dataframe:')
    #     print(_Fotos)
        
       

    #     final_Data =  cidades_support.join(_Fotos)
    #      #Esvaziando Dataframes
    #     cidades_support = cidades_support.iloc[0:0]
    #     _Fotos =_Fotos.iloc[0:0]
    #     print('Final_Data:')

    #     print(final_Data)
    #     #adicionar_dados_ultima_linha('raspagem.xlsx',final_Data)
    #     adicionar_dados_dataframe('raspagem.xlsx', final_Data)
    #     lista_fotos_final=[]
    #     #df = df.iloc[0:0]
     




#Ultima atualização da função TentarPegarElementoMoficidado-----------------------------------------------------------------------------------
def getIconGps(driver,tentativas =50,delay=0.5):

     
    """
     Essa função é responsável por verificar se o elemento de gps existe na foto
    """
    tentar= 1
    while tentar <=tentativas:
         
         try:
            #driver.find_element(By.XPATH,'/html/body/s123-app/div/main/ng-component/div/ng-component/survey123-ui-media-viewer/div/div[1]/div[2]/div[6]/div[1]')
            driver.find_element(By.CLASS_NAME,'icon-gps')
            return True
         except Exception as e:
              print(f'Não encontrou o ícone de GPS na foto, tentativas: {tentar}')
         time.sleep(delay)
         tentar+=1
    return False

def GetCityGeneral(driver,tentativas =120,delay=1):
    """ 
    Essa função é responsável por obter a cidade. Isso é feito pegando essa informação na foto. Caso essa informação
    não exista iremos obter através do gps.
    A confirmação de que a localização não existe na foto, é data através do icone do gps que aparece em cada foto.
    Logo se iterarmos por todas as fotos e não acharmos esse ícone, iremos tentar pegar a localização através
    do GPS.
    
    """


    fotos = driver.find_elements(By.CSS_SELECTOR, 'img.ir.ir-middle')
    contador =1
    while len(fotos) == 0:
            fotos = driver.find_elements(By.CSS_SELECTOR, 'img.ir.ir-middle')
            if len(fotos)== 0:
                print(f'Imagens ainda não encontrada {'.'*contador}')
                contador+=1
            if contador == 6:
                contador = 1
            time.sleep(0.2)
    print(f'Encontrou as fotos na página, quantidade de fotos: {len(fotos)}!')

    quantidade = len(fotos)#Quantidade de fotos a serem iteradas
    qtdsemicone = 0
    tentar = 0
    while tentar <= quantidade:
        try:
            """
            Essa parte do código é responsável por iterar em todas as fotos de uma determinada linha. Logo iremos iterar em todas as fotos
            em seguida iremos clicar em cada uma e verificar se o botão de gps existe. Caso existe iremos obter a localização. Caso contrário
            não existem em todas as fotos iremos obter o a cidade através da API de GPS.

            """
            time.sleep(0.5)
            img_element = driver.find_elements(By.CSS_SELECTOR, 'img.ir.ir-middle')#Vetor com todas as fotos
            img_element[tentar].click()
            time.sleep(1.2)
            """
            O código abaixo é responsável em obter as informações com respeito ao ícone de gps
            """
            if getIconGps(driver) == True:
                elements = driver.find_elements(By.CSS_SELECTOR, 'div.item > p.main-info')
                str=elements[len(elements)-1].text
                array = str.split(', ')
                print(f'Cidade obtida na imagem:{array[1]}')
                FecharJanela(driver)
                return  [array[1],qtdsemicone]
            else:
                 print('Essa imagem não possui ícone do GPS!')
                 FecharJanela(driver)
                 qtdsemicone+=1    
        except Exception as e:
            print(f'Não conseguiu encontrar as imagems na página:{e}')
        time.sleep(0.5)
        tentar+=1
    """
    Se chegarmos aqui é porque todas as imagens foram iteradas e não encontramos o icone de gps em nenhuma.
    Logo iremos tentar obter a cidade pela função GPS
    """
    return ['',qtdsemicone]
                      

def TentarPegarElementoModificadoNovo(driver, retries=120, delay=1):
    """
    This function will try to get the element every time this task occurs, if possible.
    An error may be raised, so every time these errors occur, this function will be called again.
    It will retry up to 'retries' times with a delay of 'delay' seconds between attempts.
    """
    attempt = 0
    lista_dados_parcial = []

    while attempt < retries:
        try:
            # Parse the page source with BeautifulSoup
            DataPage = BeautifulSoup(driver.page_source, 'html.parser')
           

            # Encontrar o elemento <p> que contém um <b> com o texto 'ID Fazenda'
            element = DataPage.find('p', string=lambda t: t and 'ID Fazenda' in t)

            if element is not None:
                # Encontrar o elemento <p> irmão com a classe 'answer'
                answer_element = element.find_next('p', class_='answer')
                
                if answer_element:
                    answer_text = answer_element.get_text()
                    print(f'Fazenda:{answer_text}')
                else:
                    print("Elemento irmão com a classe 'answer' não encontrado.")
            else:
                print("Elemento contendo <b>ID Fazenda</b> não encontrado.")
                

            #This code is responsible for getting and clicking on the image
            if answer_element != "":

                _cidade,qtd=GetCityGeneral(driver)
                fotosTotais = len(driver.find_elements(By.CSS_SELECTOR, 'img.ir.ir-middle'))
                if qtd == fotosTotais:
                    print('Obtendo cidade pelo GPS')
                    _cidade = find_city_from_page_source(driver)
                    if _cidade !='':
                        print(f'Cidade encontrada pelo GPS: {_cidade}!')
                        return [_cidade,answer_text]
                    else:
                        print(f'Cidade não encontrada pelo GPS!')
                else:
                    print(f'Cidade encontrada na Imagem. Cidade: {_cidade}')
                    return [_cidade,answer_text]
            
        except AttributeError as e1:
            print(f"AttributeError: {e1}")
        
        except NoSuchElementException as e2:
            print(f"NoSuchElementException: {e2}")
        
        except ElementNotInteractableException as e3:
            print(f"ElementNotInteractableException: {e3}")
        
        except ElementClickInterceptedException as e4:
            print(f"ElementClickInterceptedException: {e4}")
        
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

        
    
        attempt += 1
        print(f"Retrying... ({attempt}/{retries})")
        time.sleep(delay)
    
    raise Exception(f"Failed to find the element after {retries} attempts")   

def CorrigirLinhasVazias(driver,df):
    lista_fotos_final = []
    lista_contador_final = []
    valores_unicos = df['Contador'].unique()
    cidades_support = pd.DataFrame(columns=["Cidade", "ID Fazendas"])
    _Fotos = pd.DataFrame(columns=["Foto das Fazendas"])
    _contador = pd.DataFrame(columns=['Contador'])

    for i in range(0, len(valores_unicos)):
            # Acessar a página
        index = int(valores_unicos[i])
        driver.get(f'https://survey123.arcgis.com/surveys/7c7eb1d6f2b74421a50e01a2d9eb2667/data?extent=-80.6008,-21.9941,-13.8039,-9.7200&objectIds={(index)}')
        # Tentar obter o elemento
        try:
            _cidade, answer_text = TentarPegarElementoModificadoNovo(driver, retries=45)
            print('TryGetElement complete: Element its avaible for to get!')
        except Exception as e:
            print(f'Error during TryGetElement: {e}')
            continue

            # Obter dados da página
        DataPage = BeautifulSoup(driver.page_source, 'html.parser')
        panel = DataPage.find('question-answer')

        lista_fotos = panel.find_all('img', attrs={'class': '0 ir ir-middle'})
            
        for foto in lista_fotos:
            data_city = {_cidade: answer_text}
                #print(data_city)
            cidades_df = pd.DataFrame(list(data_city.items()), columns=["Cidade", "ID Fazendas"])
                #print(data_city.items())
            cidades_support = pd.concat([cidades_support, cidades_df], ignore_index=True)
                #print(cidades_support)
            lista_fotos_final.append(foto['src'])
            lista_contador_final.append(index)


        print(len(lista_fotos_final))

            # Criar a planilha do Excel toda vez que o programa rodar

        AddArrayDataframe(_Fotos, lista_fotos_final)
        AddArrayDataframe(_contador,lista_contador_final)
            #print(_Fotos)
        print('Fotos dataframe:',_Fotos)
            #Criando o dataframe do contador-------------------------------------------------------------------------------
        contador = {index: answer_text}
        contador_df = pd.DataFrame(list(contador.items()), columns=["Contador", "ID Fazendas"])

        final_Data = cidades_support.join(_Fotos)
        final_Data = final_Data.join(_contador)
        print('DataFrame Final')
    
        final_Data = final_Data[['Contador', 'Cidade', 'ID Fazendas', 'Foto das Fazendas']]
        print(final_Data)
        adicionar_dados_dataframe('raspagem.xlsx', final_Data)

            # Esvaziar DataFrames
        contador_df = contador_df.iloc[0:0]
        cidades_support = cidades_support.iloc[0:0]
        _Fotos = _Fotos.iloc[0:0]
        final_Data=final_Data.iloc[0:0]
        lista_contador_final = []
        lista_fotos = []
        lista_fotos_final = []
        